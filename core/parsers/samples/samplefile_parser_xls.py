from io import BytesIO
from typing import Sequence

import openpyxl

from core.parsers.samples.samplefile_parser_abstract import AbstractFileParser


class XLSFileParser(AbstractFileParser):
    tab_names: list = None

    def _trim_row_values(self, row: Sequence) -> list:
        """
        Trim trailing empty cells from the row values and return list up to last non-empty cell.
        A cell is considered empty if it's None or a string that's empty/whitespace.
        """
        last = -1
        for i, cell in enumerate(row):
            if cell is not None:
                if isinstance(cell, str):
                    if cell.strip() != "":
                        last = i
                else:
                    # non-string non-None is considered data
                    last = i
        if last == -1:
            return []
        return [cell for cell in row[:last + 1]]

    def get_tab_names(self) -> Sequence[str] | None:
        if self.tab_names:
            return self.tab_names

        try:
            workbook = openpyxl.load_workbook(self.content, read_only=True)
            self.tab_names = workbook.sheetnames
            return self.tab_names
        except Exception as e:
            raise ValueError(f"Error reading XLS file: {e}")

    def _find_header_line(self, tab_name: str) -> tuple[int, Sequence[str]] | None:
        try:
            workbook = openpyxl.load_workbook(self.content, read_only=True)
            sheet = workbook[tab_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                trimmed = self._trim_row_values(row)
                if not trimmed:
                    # skip empty rows
                    continue

                if self._is_header_row(trimmed):
                    return row_idx, [cell if cell is not None else "" for cell in trimmed]
            return None
        except Exception as e:
            raise ValueError(f"Error processing Excel file: {e}")

    def find_header_line(self, selected_tab: int = -1) -> tuple[int, Sequence[str]]:
        tab_names = self.get_tab_names()

        if selected_tab > -1:
            tab_name = tab_names[selected_tab]
            header_line = self._find_header_line(tab_name)
            if header_line:
                self.selected_tab = selected_tab
                return header_line
        else:
            for tab_idx, tab_name in enumerate(tab_names):
                header_line = self._find_header_line(tab_name)
                if header_line:
                    self.selected_tab = tab_idx
                    return header_line

        raise ValueError("No header line found in the Excel file")

    def get_column_names(self, line: int, selected_tab: int = -1) -> Sequence[str]:
        tab_names = self.get_tab_names()
        tab_name = tab_names[selected_tab] if selected_tab != -1 else tab_names[0]
        try:
            workbook = openpyxl.load_workbook(self.content, read_only=True)
            sheet = workbook[tab_name]
            rows = list(sheet.iter_rows(min_row=line, max_row=line, values_only=True))
            if not rows:
                raise IndexError
            row_values = rows[0]
            trimmed = self._trim_row_values(row_values)
            return [cell if cell is not None else "" for cell in trimmed]
        except IndexError:
            raise ValueError(f"Row {line} does not exist in the sheet '{tab_name}'.")
        except KeyError:
            raise ValueError(f"Sheet '{tab_name}' does not exist in the Excel file.")
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {e}")

