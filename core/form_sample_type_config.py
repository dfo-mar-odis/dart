import csv
from io import BytesIO

import openpyxl
from bs4 import BeautifulSoup
from crispy_forms.helper import FormHelper
from crispy_forms.layout import Layout, Field, Row, Column
from crispy_forms.utils import render_crispy_form
from django.http import HttpResponse
from django.urls import path, reverse_lazy

from django import forms


class FileConfig:
    file_type: str
    tab: int = -1
    tab_names: list = None
    header_line_number: int = 0

    def get_file_type(self):
        return self.file_type

    def get_xls_tab_names(self) -> list:

        if self.tab_names:
            return self.tab_names

        try:
            workbook = openpyxl.load_workbook(BytesIO(self.content), read_only=True)
            self.tab_names = workbook.sheetnames
            return self.tab_names
        except Exception as e:
            raise ValueError(f"Error reading XLS file: {e}")


    def get_header_line_number(self) -> int:
        if self.file_type == 'XLS':
            return self._find_header_in_excel()
        elif self.file_type in ['CSV', 'DAT']:
            return self._find_header_in_csv_or_dat()
        else:
            raise ValueError("Unsupported file type for header detection")

    def _find_header_in_excel(self) -> int:
        try:
            tab_names = self.get_xls_tab_names()
            workbook = openpyxl.load_workbook(BytesIO(self.content), read_only=True)
            sheet = workbook[tab_names[self.tab]] if self.tab != -1 else workbook.active
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if self._is_header_row(row):
                    return row_idx
            raise ValueError("No header line found in the Excel file")
        except Exception as e:
            raise ValueError(f"Error processing Excel file: {e}")

    def _find_header_in_csv_or_dat(self) -> int:
        try:
            content_stream = BytesIO(self.content)
            content_stream.seek(0)
            reader = csv.reader(content_stream.read().decode('utf-8').splitlines())
            for row_idx, row in enumerate(reader, start=1):
                if self._is_header_row(row):
                    return row_idx
            raise ValueError("No header line found in the CSV/DAT file")
        except Exception as e:
            raise ValueError(f"Error processing CSV/DAT file: {e}")

    def _is_header_row(self, row) -> bool:
        # A heuristic to determine if a row is a header row
        if not row:
            return False
        # Check if the row has a significant number of non-empty text cells
        text_count = sum(1 for cell in row if isinstance(cell, str) and cell.strip())
        # Check if the row has a reasonable number of columns (e.g., > 3)
        return text_count > len(row) / 2 and len(row) > 3


    def __init__(self, filename, content: BytesIO, tab: int = -1):
        self.tab = tab
        self.filename = filename
        self.content = content.read()

        # Determine file type based on the file extension
        extension = filename.split('.')[-1].lower()
        if extension == 'csv':
            self.file_type = 'CSV'
        elif extension in ['xls', 'xlsx']:
            self.file_type = 'XLS'
        elif extension == 'dat':
            self.file_type = 'DAT'
        else:
            raise TypeError(f'Unsupported file type: {extension}')



class FileConfigForm(forms.Form):

    file_tab = forms.ChoiceField(choices=[])
    header_line_number = forms.IntegerField()

    def __init__(self, file_config: FileConfig, *args, **kwargs):
        initial = kwargs.pop('initial', {})
        if 'header_line_number' not in initial:
            initial['header_line_number'] = file_config.get_header_line_number()

        if 'file_tab' not in initial:
            initial['file_tab'] = file_config.tab

        super().__init__(*args, **kwargs, initial=initial)

        tab_header_row = Row(
            Column(Field('header_line_number', css_class='form-control'))
        )
        self.helper = FormHelper(self)
        self.helper.form_tag = False

        self.helper.layout = Layout(
            tab_header_row
        )

        if file_config.file_type == 'XLS':
            tab_names = file_config.get_xls_tab_names()
            self.fields['file_tab'].choices = [(i, name) for i, name in enumerate(tab_names)]

            tab_attrs = {
                'hx-post': reverse_lazy('core:form_sample_type_get_headers'),
                'hx-target': '#div_id_sample_config_form_content',
                'hx-trigger': 'change',
                'hx-indicator': "#div_id_indicator_sample_config"
            }
            tab_header_row.insert(0, Column(Field('file_tab', css_class='form-control', **tab_attrs)))

def get_file_config(request):

    # only one file can be uploaded here at a time.
    file = request.FILES.get('sample_file', None)
    soup = BeautifulSoup('<div id="div_id_sample_config_form_content"></div>', 'html.parser')
    if file:
        initial={}
        if header_line_number:=request.POST.get('file_tab', -1):
            initial['header_line_number'] = int(header_line_number)

        file_config = FileConfig(file.name, file, tab=initial['header_line_number'])
        form = FileConfigForm(file_config)
        form_content = soup.find('div')

        html = render_crispy_form(form)
        form_content.append(BeautifulSoup(html, 'html.parser'))

    return HttpResponse(soup)


url_patterns = [
    path('sample_config/header/', get_file_config, name='form_sample_type_get_headers')
]